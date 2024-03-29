from openpyxl import load_workbook
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, ParseMode, message
from telegram.ext import CallbackContext
from credentials import timezone
import datetime
import pytz
import time
import os

ADMIN_ID = load_workbook("custom/token.xlsx").active["A2"].value
TIMEZONE = timezone

def open_workbook(update, context):
    try:
        wb = load_workbook(f"custom/excel_sheet.xlsx")
    except:
        message = "Upload question spreadsheet file using /addsheet!"
        context.bot.send_message(chat_id=update.effective_chat.id, text=message)
        return None
    return wb

def show_workbooks(wb, update, context):
    # Find all sheet names
    workbooks = os.listdir("C://Users//singh//OneDrive//Desktop//Telegram bots//Questions_Bot//custom//Question")
    keyboard = []
    for item in workbooks:
        filename = item.split(".")[0]
        keyboard.append([InlineKeyboardButton(filename, callback_data=f"{item}")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    context.bot.send_message(
        chat_id=update.effective_chat.id, 
        text="Choose the spreadsheet you want to post questions from:",
        reply_markup=reply_markup,
    )
    return

def show_groups(wb, query, update, context):
    sheet = wb["Groups"]
    group_list = [item.value for item in sheet["A"]]
    text = "These are the telegram groups you administer. Choose the group you want to post questions to and then select DONE.\n\n"
    text += "Groups Selected:" + " "
    selected_groups = context.user_data.get("groups")
    if not len(selected_groups):
        text += "None"
    else:
        text += ", ".join([group for group in selected_groups])
    groups = set(group_list).symmetric_difference(set(selected_groups))
    keyboard = []
    for group in groups:
        keyboard.append([InlineKeyboardButton(group, callback_data=f"group_{group}")])
    if len(selected_groups):
        keyboard.append([InlineKeyboardButton("Done.", callback_data=f"done")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    query.edit_message_text(
        text=text, 
        reply_markup=reply_markup
    )

def validate_date(update, context):
    text = update.message.text
    try:
        date = [int(x) for x in text.split("-")]
    except:
        update.message.reply_text("Invalid message!")
    # An error might occur while running questions if the date has been set in the past
    try:
        datetime.datetime(date[0], date[1], date[2])
    except:
        update.message.reply_text("Wrong format, use YYYY-MM-DD instead.")
        return None
    context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="What time do you want to post the first question (in HH:MM:SS format)?"
    )
    return text

def validate_time(update, context):
    text = update.message.text
    timeformat = "%H:%M:%S"
    try:
        # Check if 0 problem happens here!
        datetime.datetime.strptime(text, timeformat)
    except:
        update.message.reply_text("Wrong format, use HH:MM:SS instead.")
        return None
    return text

def save_data(update, context):
    # Update workbook
    wb_main = open_workbook(update, context)
    if not wb_main:
        return
    wb_attendance = load_workbook("custom/attendance_sheet.xlsx")
    if fill_database(wb_main, wb_attendance, context):
        context.bot.send_message(chat_id=update.effective_chat.id, text="Data updated!")
    else:
        context.bot.send_message("Oops, something went wrong, try again!")

def fill_database(wb_main, wb_attendance, context):
    job_count = wb_main["count"]["B1"].value + 1
    wb_main["count"].cell(row=1, column=2, value=job_count)

    # Modify attendance sheet
    group_with_column = add_attendance_columns(wb_attendance, context)
    group_string = []
    for item in group_with_column:
        group_string.append(":".join(item))
    group_string = ", ".join(group_string)
    schedule = wb_main["Schedule"]
    data = [
        context.user_data["sheet"],
        group_string,
        context.user_data["date"],
        context.user_data["time"],
        job_count,
        context.user_data["job_type"],
        3, # question row counter
        context.user_data["workbook"],
    ]
    schedule.append(data)
    wb_main.save(filename="custom/excel_sheet.xlsx")
    
    return wb_main

def add_attendance_columns(wb, context):
    groups = context.user_data["groups"]
    raw_date = context.user_data["date"]
    date_list = [int(x) for x in raw_date.split("-")]
    date_object = datetime.datetime(date_list[0], date_list[1], date_list[2])
    weekday = date_object.strftime("%A")[0:3]
    month = f'{date_object.strftime("%B")[0:3]}.'
    remaining = f"{date_object.day}/{date_object.year}"
    date = " ".join([weekday, month, remaining])
    
    group_with_column = []
    for group in groups:
        ws = wb[group]
        headings = ["Date", "Answered", "QA"]
        max_column = ws.max_column
        for i, heading in enumerate(headings, start=1):
            ws.cell(row=1, column=max_column + i, value=heading)

        # Add schedule's date in all rows.
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=max_column + 1, value=date)
            ws.cell(row=row, column=max_column + 3, value="0")
        
        group_with_column.append([group, str(max_column + 2)])

    wb.save(filename="custom/attendance_sheet.xlsx")
    return group_with_column
        

def create_datetime(row, date_index, time_index):
    date = [int(x) for x in row[date_index].value.split("-")]
    time = [int(x) for x in row[time_index].value.split(":")]
    date_time = pytz.timezone(TIMEZONE).localize(datetime.datetime(date[0], date[1], date[2], time[0], time[1], time[2]))

    return date_time

def group_ids_by_title(wb, group_list):
    titles = [x.split(":")[0] for x in group_list]
    sheet = wb["Groups"]
    group_ids = []
    for row in sheet.iter_rows():
        if row[0].value in titles and row[1].value not in group_ids:
            group_ids.append(row[1].value)
    return group_ids

def send_message_to_ids(bot, ids, message):
    for id in ids:
        bot.send_message(chat_id=id, text=f"<b>{message}</b>", parse_mode=ParseMode.HTML)

def check_admin(update, context):
    if update.message.from_user.id != ADMIN_ID:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Sorry, I'm not allowed to talk to you!")
        return False
    return True

def extract_datetime(date_time):
    # tzinfo already set by telegram(I think)
    date = date_time.strftime("%Y-%m-%d")
    time = date_time.strftime("%H:%M:%S")

    return (date, time)

def in_run_time(wb, date_time):
    # date+time should be after date, time in schedule(as schedule is deleted after completion)
    # if after, return True, if before return False
    schedule = wb["Schedule"]
    schedule_list = []
    for row in schedule.iter_rows():
        # Here we are working on the logic that schedules get deleted after they have been executed,
        # so only time that is going to be less than current time is the time of a schedule that is currently running.
        # Instead of bool, this function can return what schedule it's currently running.
        if create_datetime(row, 2, 3) < date_time:
            schedule_list.append(
                {
                    "schedule_number": row[0].row,
                    "sheet":row[0].value, "groups":row[1].value, 
                    "date":row[2].value, "time":row[3].value,
                    "cache_question":row[6].value,
                    "workbook":row[7].value,
                }
            )
    if len(schedule_list) == 1:
        return schedule_list[0]
    elif len(schedule_list) > 1:
        print("you can't run two schedules at the same time!")
        return None
    else:
        return None

def collect_garbage():
    # Collect garbage older schedules
    try:
        wb = load_workbook("custom/excel_sheet.xlsx")
    except:
        return
    schedule = wb["Schedule"]
    for row in schedule.iter_rows():
        if create_datetime(row, 2, 3) < datetime.datetime.now(pytz.timezone(TIMEZONE)):
            schedule.delete_rows(row[0].row)
    wb.save(filename="custom/excel_sheet.xlsx")

# def send_report(wb, context, chat_history, questions, group_ids, session_start):
#     # Makes and sends reports
#     report = {}
#     for group_id in group_ids:
#         report[group_id] = {}
#         for row_history in chat_history.iter_rows():
#             if row_history[2].value == group_id:
#                 user_id = row_history[3].value
#                 time_limit = session_start # Initialize time limit for questions
#                 for row_question in questions.iter_rows(min_row=3):
#                     question = row_question[0].value
#                     response_time = create_datetime(row_history, 0, 1)
#                     report[group_id]["session_datetime"] = response_time
#                     if not report[group_id].get("response"):
#                         report[group_id]["response"] = {}
#                     time_limit += datetime.timedelta(minutes=row_question[2].value)
#                     if session_start < response_time <= time_limit:
#                         if report[group_id]["response"].get(user_id):
#                             report[group_id]["response"][user_id].append(question)
#                         else:
#                             report[group_id]["response"][user_id] = [question]
#                         break # Don't want to iterate through next questions
#     create_and_send_msg(wb,context, report)

# def create_and_send_msg(wb, context, report):
#     for group in report:
#         message = f"{get_group_name_by_id(wb, group)}\n"
#         date, time = extract_datetime(report[group]["session_datetime"])
#         message += f"{date} {time}\n\n"
#         for user in report[group]["response"]:
#             # user is user_id
#             message += f"{get_user_name(context, group, user)}: "
#             question_list = report[group]["response"][user]
#             message += ", ".join(set([str(x) for x in question_list])) + "\n"
            
#         # group is group_id
#         admins = context.bot.get_chat_administrators(group)
#         # WRITE HTML WHILE SENDING
#         for admin in admins:
#             try:
#                 context.bot.send_message(chat_id=admin.user.id, text=message, parse_mode=ParseMode.HTML)
#             except:
#                 continue

def get_user_name(context, group_id, user_id):
    user = context.bot.get_chat_member(chat_id=group_id, user_id=user_id).user
    first_name = user.first_name
    last_name = user.last_name
    if not user.last_name:
        last_name = ""
    return first_name + " " + last_name

def get_group_name_by_id(wb, group_id):
    # Takes group_sheet, can change if needed in future
    for row in wb["Groups"].iter_rows():
        if row[1].value == group_id:
            return row[0].value

def get_question_number(question_sheet, question):
    for row in question_sheet.iter_rows(min_row=3):
        if row[0].value == 0:
            continue
        if row[1].value == question:
            return str(int(float(row[0].value)))
    return None

def get_question_from_number(question_sheet, number):
    for row in question_sheet.iter_rows(min_row=3):
        if row[0].row == number:
            return row[1].value

def test_auto(context: CallbackContext):
    # Runs the full list of questions according to time
    # Deletes the row from the sheet.
    job = context.job
    bot = context.bot
    job_index = job.context
    session_start = datetime.datetime.now(pytz.timezone(TIMEZONE))
    wb = load_workbook("custom/excel_sheet.xlsx")
    schedule = wb["Schedule"]
    for row in schedule.iter_rows():
        if row[4].value == job_index:
            question_workbook = row[7].value
            questions_sheet_title = row[0].value
            group_list = row[1].value.strip(", ")
            if not isinstance(group_list, list):
                group_list = [group_list]
    
    wb_question = load_workbook(f"custom/Question/{question_workbook}")
    questions = wb_question[questions_sheet_title]
    group_ids = group_ids_by_title(wb, group_list)

    # Send the question and sleep for the time limit
    for row in questions.iter_rows(min_row=3):
        if None in [row[1].value, row[2].value]:
            break
        send_message_to_ids(bot, group_ids, row[1].value)
        time.sleep(float(row[2].value) * 60)
    send_message_to_ids(bot, group_ids, message="time's up!")

    # Delete the schedule from the database
    for row in schedule.iter_rows():
        if row[4].value == job_index:
            schedule.delete_rows(row[0].row)
    wb.save(filename="custom/excel_sheet.xlsx")

def test_manual(context: CallbackContext):
    # Just posts the first question and initialises the question index.
    job = context.job
    bot = context.bot
    job_index = job.context
    wb = load_workbook("custom/excel_sheet.xlsx")
    schedule = wb["Schedule"]
    for row in schedule.iter_rows():
        if row[4].value == job_index:
            schedule_id = row[0].row
            question_workbook = row[7].value
            questions_sheet_title = row[0].value
            group_list = row[1].value.strip(", ")
            if not isinstance(group_list, list):
                group_list = [group_list]
    
    wb_question = load_workbook(f"custom/Question/{question_workbook}")
    questions = wb_question[questions_sheet_title]
    group_ids = group_ids_by_title(wb, group_list)

    # Send the first question and update the cache.
    send_message_to_ids(bot, group_ids, questions[3][1].value)
    wb["Schedule"][f"G{schedule_id}"] = 3
    wb.save(filename="custom/excel_sheet.xlsx")
    
# new change
def check_if_question_already_exists(string, question):
    # Checks if the number already exists in a string
    str_nums = string.split(", ")
    print(string)
    print(question)
    if question in str_nums:
        return True
    return False