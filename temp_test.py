from openpyxl import load_workbook

wb = load_workbook("custom/attendance_sheet.xlsx")
attendance_sheet = wb["okokokok"]
question = "Who is the greatest player in the history of the game?"
username = "raghu_palash"

def store_question(question_number, column_answered):
    column_QA = column_answered + 1
    for row in attendance_sheet.iter_rows(min_row=2):
        print(row)
        if row[0].value == username:
            stored = row[column_answered].value # A string like 1, 2, 3, 4 or None
            if not stored:
                attendance_sheet.cell(row=row[0].row, column=column_answered, value=question_number)
            else:
                attendance_sheet.cell(row=row[0].row, column=column_answered, value=", ".join([str(stored), question_number]))
            attendance_sheet.cell(row=row[0].row, column=column_QA, value=str(int(row[column_QA].value) + 1))

store_question(1, 4)
store_question(2, 4)