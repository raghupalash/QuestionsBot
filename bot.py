# We are going to use this file for main bot logic and holding data.
# For sending messages and storing data, we will use utils.py
from openpyxl.workbook import workbook
from telegram import Update
from telegram.ext import (
    Updater, 
    CommandHandler, 
    ConversationHandler,
    MessageHandler,
    CallbackQueryHandler, 
    CallbackContext,
    Filters,
)
import logging
from openpyxl import load_workbook
from telegram.inline.inlinekeyboardbutton import InlineKeyboardButton
from telegram.inline.inlinekeyboardmarkup import InlineKeyboardMarkup
from utils import (
    TIMEZONE,
    check_if_question_already_exists, # new change
    get_question_from_number,
    test_auto,
    test_manual,
    check_admin,
    show_workbooks, 
    show_groups, 
    open_workbook, 
    validate_date, 
    validate_time, 
    save_data, 
    create_datetime,
    in_run_time,
    collect_garbage,
    get_question_number
)
from credentials import timezone
import time
import datetime
import pytz
from test_utlis import start_admin

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)

TIMEZONE = timezone
  
ONE, TWO, THREE, FOUR, FIVE = range(5)


def add_sheet_options(update, context):
    if not check_admin(update, context):
        ConversationHandler.END
    # in open workbook error message, say "use /addsheet to add sheet"
    keyboard = [
        [InlineKeyboardButton("questions sheet", callback_data="question_sheet")],
        [InlineKeyboardButton("attendance sheet", callback_data="attendance_sheet")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    message = "Which sheet do you want to add?"
    update.message.reply_text(text=message, reply_markup=reply_markup)
    return ONE

def ask_for_sheet(update, context):
    query = update.callback_query
    query.answer()
    context.user_data["sheet_type"] = query.data
    message = f"Alright, send me the {query.data.split('_')[0]} sheet:"
    query.edit_message_text(text=message)
    return TWO

def incoming_document(update, context):
    document = update.message.document
    mime_type = document.mime_type
    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        # writing to a custom file
        if context.user_data["sheet_type"] == "question_sheet":
            file_location = f"custom/Question/{document.file_name}"
        else:
            file_location = f"custom/attendance_sheet.xlsx"
        with open(file_location, 'wb') as f:
            file = context.bot.get_file(update.message.document)
            file.download(out=f)
        
        context.bot.send_message(chat_id=update.effective_chat.id, 
            text="Sheet uploaded, press /start to view sheets from.")
    else:
        context.bot.send_message(chat_id=update.effective_chat.id,
            text="Only ms office spreadsheet allowed with extension '.xlsx'")
    return ConversationHandler.END

def start(update: Update, context:CallbackContext):
    # Allow only pms
    if update.message.chat.type == "group":
        return
    # Allow only admin
    if not check_admin(update, context):
        return ConversationHandler.END
    # Check if there's an attendance sheet at all.
    try:
        load_workbook("custom/attendance_sheet.xlsx")
    except:
        message = "Upload attendance spreadsheet file using /addsheet!"
        context.bot.send_message(chat_id=update.effective_chat.id, text=message)
        return ConversationHandler.END

    # Load worksheet
    wb = open_workbook(update, context)
    if not wb:
        return ConversationHandler.END
    show_workbooks(wb, update, context)
    return ONE

def sheets(update, context):
    query = update.callback_query
    query.answer()

    workbook = query.data
    context.user_data["workbook"] = workbook
    wb = load_workbook(f"custom/Question/{workbook}")
    keyboard = []
    for sheet in wb.worksheets:
        if sheet.title not in ["Schedule", "Groups", "count"]:
            keyboard.append([InlineKeyboardButton(str(sheet.title), callback_data=f"sheet_{sheet.title}")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    query.edit_message_text(
        text="Choose the sheet you want to post questions from:",
        reply_markup=reply_markup,
    )
    return TWO

def groups(update: Update, context: CallbackContext):
    query = update.callback_query
    query.answer()

    data = query.data.split("_")
    if data[0] == "sheet":
        context.user_data["sheet"] = data[1]
        context.user_data["groups"] = []
    elif data[0] == "group":
        # Avoid repeatition while adding into list
        if data[1] not in context.user_data.get("groups"):
            context.user_data["groups"].append(data[1])
    else:
        # User pressed "done"
        keyboard = [
            [InlineKeyboardButton("Automatic", callback_data="job_automatic")],
            [InlineKeyboardButton("Manual", callback_data="job_manual")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        query.edit_message_text(
            text="Select the job type:",
            reply_markup=reply_markup
        ) # Not a good design
        return THREE

    wb = open_workbook(update, context)
    if not wb:
        return ConversationHandler.END
    show_groups(wb, query, update, context)

def job_type(update: Update, context: CallbackContext):
    query = update.callback_query
    query.answer()
    data = query.data.split("_")
    context.user_data["job_type"] = data[1]
    query.edit_message_text(text="Write the date you want to post questions (in YYYY-MM-DD format):") # Not a good design
    return FOUR

    
def schedule_date(update: Update, context: CallbackContext):
    date = validate_date(update, context)
    if not date:
        return FOUR
    context.user_data["date"] = date
    return FIVE

def schedule_time(update: Update, context: CallbackContext):
    time = validate_time(update, context)
    if not time:
        return FIVE
    context.user_data["time"] = time
    save_data(update, context)
    return ConversationHandler.END

def add_group(update: Update, context: CallbackContext):
    if update.message.chat.type == "private":
        return
    if update.message.from_user.id != ADMIN_ID:
        update.message.reply_text("Sorry, only the bot owner can add the bot in the database.")
        return
    # Save group info
    group_title = update.message.chat.title
    group_id = update.message.chat.id
    wb = open_workbook(update, context)
    if not wb:
        return
    sheet = wb["Groups"]
    for item in sheet['B']:
        if item.value == group_id:
            update.message.reply_text("The group is already in the database")
            return
    data = [group_title, group_id]
    sheet.append(data)
    wb.save(filename="custom/excel_sheet.xlsx")
    update.message.reply_text("Group added to the database.")
    return ConversationHandler.END

def cancel(update, context):
    ConversationHandler.END

def set_jobs(update: Update, context: CallbackContext):
    if update.message.chat.type == "group":
        return
    if not check_admin(update, context):
        return
    sheet = open_workbook(update, context)["Schedule"]
    for row in sheet.iter_rows():
        time = create_datetime(row, 2, 3)
        if context.user_data["job_type"] == "automatic":
            context.job_queue.run_once(test_auto, time, context=row[4].value, name=f"job{row[0].row}")
        else:
            context.job_queue.run_once(test_manual, time, context=row[4].value, name=f"job{row[0].row}")
    update.message.reply_text("You are all set!!! The questions are scheduled to run at the date and time you have chosen.")

def handle_user_responses(update: Update, context:CallbackContext):
    message = update.message
    if not message:
        message = update.edited_message
    if message.chat.type != "group":
        return
    # Check if message is a reply to another message.
    reply_to = message.reply_to_message
    if not reply_to:
        return

    current_group = message.chat.title
    wb_main = load_workbook("custom/excel_sheet.xlsx")
    wb_attendance = load_workbook("custom/attendance_sheet.xlsx")
    attendance_sheet = wb_attendance[current_group]

    date_time = datetime.datetime.now(pytz.timezone(TIMEZONE))
    schedule = in_run_time(wb_main, date_time)
    # Check if we currently are in a schedule
    if not schedule:
        return

    wb_question = load_workbook(f"custom/Question/{schedule['workbook']}")
    question_sheet = wb_question[schedule["sheet"]]
    question_number = get_question_number(question_sheet, reply_to.text)
    if not question_number:
        return

    groups = group_and_report_column(schedule["groups"])
    # This column will be affected in the group
    column_answered = groups[current_group]
    column_QA = column_answered + 1

    # Add these questions in attendance sheet.
    # Here we operate on the understanding that the schedule is deleted after it has been executed.
    for row in attendance_sheet.iter_rows(min_row=2):
        if row[0].value == message.from_user.username:
            stored = row[column_answered - 1].value # A string like 1, 2, 3, 4 or None
            question_count = 0
            if not stored:
                attendance_sheet.cell(row=row[0].row, column=column_answered, value=question_number)
                question_count = 1
            else:
                if not check_if_question_already_exists(stored, question_number):
                    attendance_sheet.cell(row=row[0].row, column=column_answered, value=", ".join([str(stored), question_number]))
                    question_count = 1
            attendance_sheet.cell(row=row[0].row, column=column_QA, value=str(int(row[column_QA - 1].value) + question_count))

    wb_attendance.save(filename="custom/attendance_sheet.xlsx")

def send_next_question(update, context):
    # Check if this is a group and if the command is being sent by a group admin.
    if update.message.chat.type != "group":
        context.bot.send_message("Use this command in a group where schedule is running.")
        return
    if not check_if_group_admin(update, context.bot):
        update.message.reply_text("Only admins allowed to use this method.")
        return

    wb = open_workbook(update, context)
    # Check if we are in a running schedule for this group.
    date_time = datetime.datetime.now(pytz.timezone(TIMEZONE))
    schedule = in_run_time(wb, date_time) # Operates on the understanding that one schedule run at a time.
    if not schedule:
        return

    # Get the cached question number. (on 4th col we have manual/auto and 5th we have the question cache.)
    # Also increase the value in the cache
    question_number = schedule["cache_question"] + 1
    wb_question = load_workbook(f'custom/Question/{schedule["workbook"]}')
    question = get_question_from_number(wb_question[schedule["sheet"]], question_number)
    if not question:
        # End of schedule.
        context.bot.send_message(chat_id=update.effective_chat.id, text="time's up!")

        # Delete the schedule from the database
        wb["Schedule"].delete_rows(schedule["schedule_number"])
        wb.save(filename="custom/excel_sheet.xlsx")
        return
    wb["Schedule"][f"G{schedule['schedule_number']}"] = question_number
    wb.save(filename="custom/excel_sheet.xlsx")
    
    # Send the next question.
    context.bot.send_message(chat_id=update.effective_chat.id, text=question)
    
    
    
def check_if_group_admin(update, bot):
    admin_objects = bot.get_chat_administrators(update.message.chat.id)
    current_user_id = update.message.from_user.id
    for admin in admin_objects:
        if current_user_id == admin.user.id:
            return True
    return False


def group_and_report_column(groups):
    # Takes a list of groups combined with column to be edited and returns a dict.
    if not isinstance(groups, list):
        groups = [groups]
    data = {}
    for group in groups:
        split = group.split(":")
        data[split[0]] = int(split[1])
    return data

def help(update, context):
    message = (
        "/start - Add a schedule.(pm only)\n"
        "/addsheet - Add a sheet.(pm only)\n"
        "/add - Add group to the database.(group chat only)\n"
        "/set - Run this everytime you run the bot, and everytime you make a schedule " 
        "this sets the schedules inside a job."
    )
    context.bot.send_message(chat_id=update.effective_chat.id, text=message)

def main():
    print("Hello World!")
    wb_token = load_workbook("custom/token.xlsx")
    token = wb_token.active["A1"].value
    global ADMIN_ID
    ADMIN_ID = wb_token.active["A2"].value

    updater = Updater(token=token)
    
    dispatcher = updater.dispatcher

    # Garbage collect older schedules
    collect_garbage()

    start_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            ONE: [CallbackQueryHandler(sheets)],
            TWO: [CallbackQueryHandler(groups)],
            THREE: [CallbackQueryHandler(job_type)],
            FOUR: [MessageHandler(Filters.text, schedule_date)],
            FIVE: [MessageHandler(Filters.text, schedule_time)],
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )

    document_handler = ConversationHandler(
        entry_points=[CommandHandler("addsheet", add_sheet_options)],
        states={
            ONE: [CallbackQueryHandler(ask_for_sheet)],
            TWO: [MessageHandler(Filters.document, incoming_document)]
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )
    dispatcher.add_handler(CommandHandler("start_admin", start_admin))
    dispatcher.add_handler(CommandHandler("next", send_next_question))
    dispatcher.add_handler(document_handler)
    dispatcher.add_handler(CommandHandler("add", add_group))
    dispatcher.add_handler(CommandHandler("set", set_jobs))
    dispatcher.add_handler(CommandHandler("help", help))
    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(MessageHandler(Filters.text, handle_user_responses))
    

    updater.start_polling()

    updater.idle()

if __name__ == "__main__":
    main()