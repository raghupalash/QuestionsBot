from openpyxl import load_workbook

wb = load_workbook("custom/attendance_sheet.xlsx")
ws = wb["okokokok"]

ws.cell(row=1, column=5, value=0)
ws.cell(row=2, column=5, value=0)
wb.save(filename="custom/attendance_sheet.xlsx")