from openpyxl import load_workbook
from utils import TIMEZONE, create_datetime, fill_database, open_workbook, show_sheets, test_auto, test_manual
from telegram.ext import ConversationHandler
from datetime import datetime, timedelta
from credentials import timezone
import pytz

TIMEZONE = timezone
"""
These functions are implemented for making testing the bot easier for us.
start_admin_manual for manual mode, and start_admin_auto for automatic.

What's going to happend is, I will give a special commnand to the bot in the pm, and it will
make a schedule for a given sheet, given group, time = time.now() + 10 seconds. Then it's going to set the schedule.
"""
def start_admin(update, context):
    # Upload a new worksheet
    # Load worksheet
    context.user_data["sheet"] = "Basketball"
    context.user_data["groups"] = ["okokokok"]
    context.user_data["date"] = datetime.today().strftime('%Y-%m-%d')
    time = datetime.now() + timedelta(seconds=5)
    time = pytz.timezone(TIMEZONE).localize(time)
    context.user_data["time"] = time.strftime("%H:%M:%S")
    context.user_data["job_type"] = "manual"

    wb_main = open_workbook(update, context)
    wb_attendance = load_workbook("custom/attendance_sheet.xlsx")

    wb_main["Schedule"].delete_rows(1, 100)
    wb_main.save("custom/excel_sheet.xlsx")

    fill_database(wb_main, wb_attendance, context)

    # Set the schedule
    sheet = wb_main["Schedule"]
    for row in sheet.iter_rows():
        
        if context.user_data["job_type"] == "automatic":
            context.job_queue.run_once(test_auto, time, context=row[4].value, name=f"job{row[0].row}")
        else:
            context.job_queue.run_once(test_manual, time, context=row[4].value, name=f"job{row[0].row}")
    update.message.reply_text("You are all set!!! The questions are scheduled to run at the date and time you have chosen.")
